import csv
import pandas as pd
from datetime import datetime
import os
import calendar
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def get_template():
    wb = load_workbook(filename = 'template/xangle_template_unlockschedule.xlsx')
    return wb


def is_dir_else_create(project_name):
    if os.path.isdir(project_name) is False:
        os.mkdir(project_name)


def first_of_month(date):
    date_obj = datetime.strptime(date, '%Y/%m/%d')

    # Format the date object to the first day of the month
    first_day_date = date_obj.replace(day=1)
    return first_day_date

def last_of_month(date):
    date_obj = datetime.strptime(date, '%Y/%m/%d')
    last_day = calendar.monthrange(date_obj.year, date_obj.month)[1]
    last_day_date = date_obj.replace(day=last_day)
    return last_day_date

def default_month(date):
    default_date = datetime.strptime(date, '%Y/%m/%d')
    return default_date



def extractor(project_name, dateoption):
    df_dict = {}
    df = pd.read_csv(f"input_csvs/{project_name}.csv")

    if dateoption == "1":
        date_revised = [default_month(day) for day in df.columns[1:]]
    elif dateoption == "2":
        date_revised = [first_of_month(day) for day in df.columns[1:]]
    else:
        date_revised = [last_of_month(day) for day in df.columns[1:]]

    if date_revised[0] == date_revised[1]:
        date_revised = date_revised[1:]
        df = df.drop(columns = df.columns[1])
    for i,column in enumerate(df.values):
        col_string = [str(num_string) for num_string in column[1:]]
        data = {
            '락업해제일자* (yyyy-mm-dd)' : date_revised,
            '락업해제일자': [int(num_string.replace(',', '')) if num_string.lower() != 'nan' else 0 for num_string in col_string]

        }
        df = pd.DataFrame(data)
        df_dict[column[0]] = data
    return df_dict

def excel_writer(project_name, k, v, dateoption):
    dateoption_str = {"1": "default", "2": "first", "3": "last"}

    wb = get_template()
    ws = wb['Sheet1']
    v = pd.DataFrame(v)
    rows, cols = v.shape

    for r in range(rows):
        for c in range(cols):
            cell = ws.cell(row=r+2, column=c+1)
            cell.value = v.iloc[r,c]

    filename = f'output_csvs/{project_name}/{k}_extracted_{dateoption_str[dateoption]}.xlsx'
    wb.save(filename)
    print(f"{filename} saved!")

def main():
    if not os.path.isdir("input_csvs"):
        os.mkdir("input_csvs")
    if not os.path.isdir("output_csvs"):
        os.mkdir("output_csvs")
    project_name = input("which project would you like to extract?...")
    is_dir_else_create('output_csvs/' + project_name)
    while True:
        dateoption = input("""which date format do you wish to apply?...
    [1] default (as-is) [2] first day of the month [3] last day of the month...""")
        if dateoption in ["1", "2", "3"]:
            break
        else:
            print("please answer with ONE NUMBER between 1 to 3!...")
            pass
    df_dict = extractor(project_name, dateoption)
    for k,v in df_dict.items():
        excel_writer(project_name, k, v, dateoption)
    print(f"{project_name} extraction finished!")

main()
